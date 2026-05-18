"""Load raw Excel files from data/inputs/ into a standardised DataFrame.

Handles the active cohort workbook (QTX_AX nov 2025) using openpyxl
to navigate the non-standard header layout (super-headers in row 3,
field headers in row 4, data from row 5).

Formula residue past the last real patient row is discarded by stopping
at the last row that has a non-null S/N value.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from qtx.utils.config import get_path, get_schema_config
from qtx.utils.logging import get_logger

log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Column name → canonical mapping
# ---------------------------------------------------------------------------

# Maps raw header text (lower-stripped) to canonical names from schema.yaml.
# Duplicated generic names (e.g. "Post - Pre Gains (s)") are resolved by
# position at runtime; this dict handles the unambiguous ones.
_RAW_TO_CANONICAL: dict[str, str] = {
    "s/n": "sn",
    "name": "name",
    "gender": "gender",
    "age": "age",
    "joined with pain?": "joined_with_pain",
    "pain improved?": "pain_improved",
    "pre bixeps vas": "pre_vas",
    "post bixeps vas": "post_vas",
    "vas change": "vas_change",
    "tags": "tags",
    "location of pain if any \n(ll vs body)": "pain_location",
    # TUG block (col 13-16)
    "pre tug result (s)": "pre_tug_s",
    "post tug result (s)": "post_tug_s",
    # 5XSST block (col 18-21)
    "pre 5xsst result (s)": "pre_5xsst_s",
    "post 5xsst result (s)": "post_5xsst_s",
    # Normal gait block (col 23-28)
    "pre normal time (s)": "pre_normal_time_s",
    "post normal time (s)": "post_normal_time_s",
    "pre normal gs (m/s)": "pre_normal_gs_ms",
    "post normal gs (m/s)": "post_normal_gs_ms",
    # Fast gait block (col 30-35)
    "pre fast time (s)": "pre_fast_time_s",
    "post fast time (s)": "post_fast_time_s",
    "pre fast gait speed (m/s)": "pre_fast_gs_ms",
    "post fast gait speed (m/s)": "post_fast_gs_ms",
    # SPPB block (col 37-40)
    "baseline sppb": "baseline_sppb",
    "post sppb": "post_sppb",
    "sppb diff": "sppb_change",
    "actual or estimated sppb": "sppb_source",
    # Dosing / phone
    "once/ l + r 10/both legs": "usage_frequency",
    "phone number": "phone_number",
}

# Positional disambiguation for repeated generic names.
# Key: 0-based column index → canonical name
_POSITIONAL_OVERRIDES: dict[int, str] = {
    15: "tug_change_s",       # "Post - Pre Gains (s)" in TUG block
    16: "tug_change_pct",     # "Change over Pre (%)" in TUG block
    20: "sst_change_s",       # "Post - Pre Gains (s)" in 5XSST block
    21: "sst_change_pct",     # "Change over Pre (%)" in 5XSST block
    27: "normal_gs_change_ms",  # "Post - Pre Gains (m/s)" in Normal gait
    28: "normal_gs_change_pct",  # "Change over Pre (%)" in Normal gait
    34: "fast_gs_change_ms",    # "Post - Pre Gains (m/s)" in Fast gait
    35: "fast_gs_change_pct",   # "Change over Pre (%)" in Fast gait
}


def _build_column_names(header_row: tuple[Any, ...]) -> list[str]:
    """Map raw header values to canonical names, falling back to cleaned raw text."""
    col_names: list[str] = []
    seen: dict[str, int] = {}

    for col_idx, raw_val in enumerate(header_row):
        # Positional override takes priority
        if col_idx in _POSITIONAL_OVERRIDES:
            col_names.append(_POSITIONAL_OVERRIDES[col_idx])
            continue

        if raw_val is None:
            # Unnamed column — generate a placeholder
            name = f"_col_{col_idx}"
        else:
            cleaned = str(raw_val).strip()
            canonical = _RAW_TO_CANONICAL.get(cleaned.lower())
            if canonical:
                name = canonical
            else:
                # Sanitise for use as a DataFrame column name
                name = cleaned.lower().replace(" ", "_").replace("/", "_").replace("\n", "_")

        # Deduplicate
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0

        col_names.append(name)

    return col_names


def load_raw(path: str | Path | None = None) -> pd.DataFrame:
    """Read the source Excel with openpyxl and return a raw DataFrame.

    Parameters
    ----------
    path:
        Explicit path to the Excel workbook.  When *None* the path is read
        from ``settings.yaml`` under ``paths.raw_excel``.

    Returns
    -------
    pd.DataFrame
        One row per patient, columns use canonical names where available.
        Raw (messy) values are preserved — no coercion is applied here.
    """
    if path is None:
        path = get_path("raw_excel")
    path = Path(path)

    log.info("Opening workbook: %s", path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Use the single sheet present
    ws = wb.active
    log.info("Reading sheet: %s", ws.title)

    # -----------------------------------------------------------------------
    # Build column name list from row 4
    # -----------------------------------------------------------------------
    header_row_values: tuple[Any, ...] = next(
        ws.iter_rows(min_row=4, max_row=4, values_only=True)
    )
    col_names = _build_column_names(header_row_values)
    n_cols = len(col_names)
    log.debug("Detected %d columns; names: %s", n_cols, col_names)

    # -----------------------------------------------------------------------
    # Read patient data rows (row 5+), stopping at last non-null S/N
    # -----------------------------------------------------------------------
    # S/N is at column index 1 (0-based) in the row tuple
    sn_col_idx = 1

    rows: list[list[Any]] = []
    last_real_row_num = 0

    for row_num, row_vals in enumerate(
        ws.iter_rows(min_row=5, values_only=True), start=5
    ):
        sn_val = row_vals[sn_col_idx] if len(row_vals) > sn_col_idx else None
        if sn_val is not None:
            # Pad / trim row to expected column count
            row_list = list(row_vals)
            if len(row_list) < n_cols:
                row_list.extend([None] * (n_cols - len(row_list)))
            else:
                row_list = row_list[:n_cols]
            rows.append(row_list)
            last_real_row_num = row_num

    wb.close()

    df = pd.DataFrame(rows, columns=col_names)
    log.info(
        "Loaded %d patient rows, %d columns (last worksheet row: %d)",
        len(df),
        len(df.columns),
        last_real_row_num,
    )
    return df
